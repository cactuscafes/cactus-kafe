#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
e-Arşiv / e-Fatura ayıklayıcı
=============================

Verilen klasördeki e-postaları ve fatura eklerini tarar, UBL-TR (e-Arşiv /
e-Fatura) XML'lerini ve GİB e-Arşiv PDF'lerini ayrıştırıp tek bir tabloya
döker.

Desteklenen girdiler (iç içe olanlar dahil):
  .eml / .msg-benzeri RFC822  -> ekleri açılır
  .zip                        -> içeriği özyinelemeli açılır
  .xml                        -> UBL-TR Invoice olarak ayrıştırılır
  .pdf                        -> GİB e-Arşiv PDF metni regex ile ayrıştırılır

Kullanım:
  python3 earsiv_ayikla.py <klasor_veya_dosya> [...] [-o cikti_adi] [--vkn 1234567890]

Çıktılar: <cikti_adi>.xlsx, <cikti_adi>.csv (Excel-TR uyumlu, UTF-8 BOM)
"""

from __future__ import annotations

import argparse
import csv
import email
import email.policy
import io
import os
import re
import sys
import zipfile
from datetime import datetime
from xml.etree import ElementTree as ET

# ---------------------------------------------------------------- yardımcılar

# UBL-TR tutarları "1234.56", GİB PDF'leri "1.234,56" biçiminde yazar.
_SAYI = re.compile(r"-?[\d.,]+")


def sayi(deger, tr_bicim=False):
    """Metni ondalık sayıya çevirir; çevrilemezse None döner."""
    if deger is None:
        return None
    m = _SAYI.search(str(deger).strip())
    if not m:
        return None
    s = m.group(0)
    if tr_bicim or ("," in s and s.rfind(",") > s.rfind(".")):
        s = s.replace(".", "").replace(",", ".")   # 1.234,56 -> 1234.56
    else:
        s = s.replace(",", "")                     # 1,234.56 -> 1234.56
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def yerel_ad(etiket: str) -> str:
    """'{ns}Invoice' -> 'Invoice'"""
    return etiket.rsplit("}", 1)[-1]


def bul(kok, *yol):
    """Ad alanından bağımsız tek düğüm arar: bul(kok, 'TaxTotal', 'TaxAmount')"""
    dugum = kok
    for ad in yol:
        sonraki = None
        for cocuk in dugum:
            if yerel_ad(cocuk.tag) == ad:
                sonraki = cocuk
                break
        if sonraki is None:
            return None
        dugum = sonraki
    return dugum


def metin(kok, *yol, varsayilan=None):
    d = bul(kok, *yol)
    if d is None or d.text is None:
        return varsayilan
    return d.text.strip() or varsayilan


def cocuklar(kok, ad):
    return [c for c in kok if yerel_ad(c.tag) == ad]


# ------------------------------------------------------- girdi dosyası açıcısı

ARSIV_UZANTILARI = {".zip"}
POSTA_UZANTILARI = {".eml", ".mbox", ".msg"}


def dosyalari_gez(yollar, _derinlik=0):
    """(kaynak_etiketi, dosya_adi, icerik_bytes) üçlüleri üretir.

    .eml ve .zip dosyaları açılıp içindekiler de üretilir (özyinelemeli).
    """
    if _derinlik > 6:                      # zip-bombasına karşı basit sınır
        return
    for yol in yollar:
        if os.path.isdir(yol):
            alt = [os.path.join(yol, a) for a in sorted(os.listdir(yol))]
            yield from dosyalari_gez(alt, _derinlik)
            continue
        if not os.path.isfile(yol):
            print(f"  ! bulunamadı: {yol}", file=sys.stderr)
            continue
        with open(yol, "rb") as f:
            veri = f.read()
        yield from _ac(os.path.basename(yol), veri, os.path.basename(yol), _derinlik)


def _ac(kaynak, veri, ad, derinlik):
    uzanti = os.path.splitext(ad)[1].lower()

    if uzanti in POSTA_UZANTILARI or veri[:5] in (b"From ", b"Retur", b"Deliv"):
        yield from _eposta_ac(kaynak, veri, derinlik)
        return

    if uzanti in ARSIV_UZANTILARI or veri[:2] == b"PK":
        yield from _zip_ac(kaynak, veri, derinlik)
        return

    yield (kaynak, ad, veri)


def _eposta_ac(kaynak, veri, derinlik):
    """E-postanın eklerini çıkarır; gövdeyi yok sayar."""
    try:
        msg = email.message_from_bytes(veri, policy=email.policy.default)
    except Exception as e:
        print(f"  ! e-posta okunamadı ({kaynak}): {e}", file=sys.stderr)
        return
    konu = (msg.get("Subject") or "").strip()
    etiket = f"{kaynak} [{konu[:60]}]" if konu else kaynak

    for parca in msg.walk():
        if parca.get_content_maintype() == "multipart":
            continue
        ad = parca.get_filename()
        if not ad:
            # Ek adı yoksa yalnızca XML/PDF gövdeleri ilgimizi çeker.
            ctype = parca.get_content_type()
            if ctype not in ("application/xml", "text/xml", "application/pdf"):
                continue
            ad = "govde.xml" if "xml" in ctype else "govde.pdf"
        try:
            icerik = parca.get_payload(decode=True)
        except Exception:
            continue
        if not icerik:
            continue
        yield from _ac(etiket, icerik, ad, derinlik + 1)


def _zip_ac(kaynak, veri, derinlik):
    try:
        zf = zipfile.ZipFile(io.BytesIO(veri))
    except Exception as e:
        print(f"  ! zip açılamadı ({kaynak}): {e}", file=sys.stderr)
        return
    for bilgi in zf.infolist():
        if bilgi.is_dir() or bilgi.filename.startswith("__MACOSX/"):
            continue
        try:
            icerik = zf.read(bilgi)
        except Exception:
            continue
        yield from _ac(kaynak, icerik, os.path.basename(bilgi.filename), derinlik + 1)


# ------------------------------------------------------------ UBL-TR ayrıştır

def taraf_bilgisi(fatura, taraf_adi):
    """AccountingSupplierParty / AccountingCustomerParty -> (unvan, vkn)"""
    taraf = bul(fatura, taraf_adi, "Party")
    if taraf is None:
        return (None, None)

    unvan = metin(taraf, "PartyName", "Name")
    if not unvan:  # şahıs şirketi: ad + soyad
        kisi = bul(taraf, "Person")
        if kisi is not None:
            parcalar = [metin(kisi, "FirstName"), metin(kisi, "FamilyName")]
            unvan = " ".join(p for p in parcalar if p) or None

    vkn = None
    for kimlik in cocuklar(taraf, "PartyIdentification"):
        d = bul(kimlik, "ID")
        if d is None or not (d.text or "").strip():
            continue
        sema = (d.attrib.get("schemeID") or "").upper()
        deger = d.text.strip()
        if sema in ("VKN", "TCKN"):
            vkn = deger
            break
        if vkn is None and deger.isdigit() and len(deger) in (10, 11):
            vkn = deger
    return (unvan, vkn)


def kdv_dokumu(fatura):
    """KDV toplamı, oran bazında döküm ve tevkifat bilgisini çıkarır."""
    kdv_toplam = 0.0
    tevkifat = 0.0
    matrah_toplam = 0.0
    satirlar = []

    for tt in cocuklar(fatura, "TaxTotal"):
        for alt in cocuklar(tt, "TaxSubtotal"):
            matrah = sayi(metin(alt, "TaxableAmount"))
            tutar = sayi(metin(alt, "TaxAmount"))
            oran = sayi(metin(alt, "Percent"))
            kod = metin(bul(alt, "TaxCategory", "TaxScheme") or ET.Element("x"),
                        "TaxTypeCode") or ""
            ad = metin(bul(alt, "TaxCategory", "TaxScheme") or ET.Element("x"),
                       "Name") or ""
            if tutar is None:
                continue
            # 0015 = KDV, 9015 = KDV tevkifatı (GİB vergi kodları)
            if kod == "9015" or "TEVKİFAT" in ad.upper() or "TEVKIFAT" in ad.upper():
                tevkifat += tutar
                continue
            if kod and kod != "0015" and "KDV" not in ad.upper():
                continue  # ÖTV, ÖİV vb. toplamlara katılmaz
            kdv_toplam += tutar
            if matrah is not None:
                matrah_toplam += matrah
            oran_txt = f"%{oran:g}" if oran is not None else "?"
            satirlar.append(f"{oran_txt}: {matrah or 0:.2f}→{tutar:.2f}")

    return (
        round(kdv_toplam, 2) if satirlar else None,
        "; ".join(satirlar) or None,
        round(tevkifat, 2) if tevkifat else None,
        round(matrah_toplam, 2) if satirlar else None,
    )


def ubl_ayristir(veri: bytes, kaynak: str, ad: str):
    try:
        kok = ET.fromstring(veri)
    except ET.ParseError:
        return None

    # İmzalı zarf ya da paket içindeki Invoice düğümünü bul.
    if yerel_ad(kok.tag) != "Invoice":
        fatura = None
        for dugum in kok.iter():
            if yerel_ad(dugum.tag) == "Invoice":
                fatura = dugum
                break
        if fatura is None:
            return None
    else:
        fatura = kok

    satici_unvan, satici_vkn = taraf_bilgisi(fatura, "AccountingSupplierParty")
    alici_unvan, alici_vkn = taraf_bilgisi(fatura, "AccountingCustomerParty")
    kdv, kdv_detay, tevkifat, kdv_matrahi = kdv_dokumu(fatura)

    parasal = bul(fatura, "LegalMonetaryTotal")
    def pm(alan):
        return sayi(metin(parasal, alan)) if parasal is not None else None

    profil = (metin(fatura, "ProfileID") or "").upper()
    return {
        "belge_turu": "e-Arşiv" if "ARSIV" in profil or "ARŞIV" in profil else "e-Fatura",
        "profil": profil or None,
        "senaryo": metin(fatura, "InvoiceTypeCode"),
        "fatura_no": metin(fatura, "ID"),
        "ettn": metin(fatura, "UUID"),
        "tarih": metin(fatura, "IssueDate"),
        "saat": metin(fatura, "IssueTime"),
        "satici_unvan": satici_unvan,
        "satici_vkn": satici_vkn,
        "alici_unvan": alici_unvan,
        "alici_vkn": alici_vkn,
        "para_birimi": metin(fatura, "DocumentCurrencyCode") or "TRY",
        "mal_hizmet_toplami": pm("LineExtensionAmount"),
        "iskonto": pm("AllowanceTotalAmount"),
        "matrah": pm("TaxExclusiveAmount") if pm("TaxExclusiveAmount") is not None else kdv_matrahi,
        "kdv": kdv,
        "kdv_detay": kdv_detay,
        "tevkifat": tevkifat,
        "vergiler_dahil": pm("TaxInclusiveAmount"),
        "odenecek": pm("PayableAmount"),
        "kaynak": kaynak,
        "dosya": ad,
        "format": "XML",
    }


# ---------------------------------------------------------------- PDF ayrıştır

# GİB e-Arşiv PDF'lerindeki standart etiketler.
PDF_DESENLERI = {
    "ettn":               r"ETTN\s*[:\-]?\s*([0-9a-fA-F\-]{30,40})",
    "fatura_no":          r"Fatura\s*No\s*[:\-]?\s*([A-Z0-9]{10,20})",
    "tarih":              r"Fatura\s*Tarihi\s*[:\-]?\s*([\d]{1,2}[./-][\d]{1,2}[./-][\d]{2,4})",
    "mal_hizmet_toplami": r"Mal\s*Hizmet\s*Toplam\s*Tutar[ıi]\s*[:\-]?\s*([\d.,]+)",
    "iskonto":            r"Toplam\s*[İI]skonto\s*[:\-]?\s*([\d.,]+)",
    "kdv":                r"Hesaplanan\s*KDV(?:\s*\(?%?\s*\d+\)?)?\s*[:\-]?\s*([\d.,]+)",
    "vergiler_dahil":     r"Vergiler\s*Dahil\s*Toplam\s*Tutar\s*[:\-]?\s*([\d.,]+)",
    "odenecek":           r"[ÖO]denecek\s*Tutar\s*[:\-]?\s*([\d.,]+)",
}
PDF_VKN = re.compile(r"(?:VKN|TCKN|Vergi\s*Kimlik\s*No)\s*[:\-]?\s*(\d{10,11})")


def pdf_metninden_alanlar(metin_govde: str):
    """GİB e-Arşiv PDF metnini alanlara çevirir (PDF ayrıştırmadan bağımsız test edilebilir)."""
    kayit = {}
    for alan, desen in PDF_DESENLERI.items():
        m = re.search(desen, metin_govde, re.IGNORECASE)
        if not m:
            continue
        ham = m.group(1)
        kayit[alan] = ham if alan in ("ettn", "fatura_no", "tarih") else sayi(ham, tr_bicim=True)

    if kayit.get("tarih"):
        for kalip in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%y"):
            try:
                kayit["tarih"] = datetime.strptime(kayit["tarih"], kalip).strftime("%Y-%m-%d")
                break
            except ValueError:
                continue

    vknler = PDF_VKN.findall(metin_govde)
    if vknler:
        kayit["satici_vkn"] = vknler[0]
        if len(vknler) > 1:
            kayit["alici_vkn"] = vknler[1]

    if kayit.get("matrah") is None:
        # Matrah PDF'te ayrı yazmaz; mal/hizmet toplamı - iskonto olarak türetilir.
        mht, isk = kayit.get("mal_hizmet_toplami"), kayit.get("iskonto") or 0
        if mht is not None:
            kayit["matrah"] = round(mht - isk, 2)
    return kayit


def pdf_ayristir(veri: bytes, kaynak: str, ad: str):
    try:
        from pypdf import PdfReader
        okuyucu = PdfReader(io.BytesIO(veri))
        govde = "\n".join((s.extract_text() or "") for s in okuyucu.pages)
    except Exception as e:
        print(f"  ! PDF okunamadı ({ad}): {e}", file=sys.stderr)
        return None

    govde = re.sub(r"[ \t]+", " ", govde)
    kayit = pdf_metninden_alanlar(govde)
    if not kayit.get("ettn") and not kayit.get("fatura_no"):
        return None       # fatura değil (ör. imza sirküleri, reklam eki)

    kayit.update({
        "belge_turu": "e-Arşiv" if re.search(r"e[\-\s]?Ar[şs]iv", govde, re.I) else "Fatura",
        "para_birimi": "TRY",
        "kaynak": kaynak,
        "dosya": ad,
        "format": "PDF",
    })
    return kayit


# --------------------------------------------------------------------- çıktı

SUTUNLAR = [
    ("tarih",              "Tarih"),
    ("fatura_no",          "Fatura No"),
    ("belge_turu",         "Belge Türü"),
    ("senaryo",            "Senaryo"),
    ("yon",                "Yön"),
    ("satici_unvan",       "Satıcı Unvan"),
    ("satici_vkn",         "Satıcı VKN/TCKN"),
    ("alici_unvan",        "Alıcı Unvan"),
    ("alici_vkn",          "Alıcı VKN/TCKN"),
    ("mal_hizmet_toplami", "Mal/Hizmet Toplamı"),
    ("iskonto",            "İskonto"),
    ("matrah",             "Matrah"),
    ("kdv",                "KDV"),
    ("kdv_detay",          "KDV Dökümü"),
    ("tevkifat",           "Tevkifat"),
    ("odenecek",           "Ödenecek Tutar"),
    ("para_birimi",        "Para Birimi"),
    ("ettn",               "ETTN"),
    ("format",             "Format"),
    ("dosya",              "Dosya"),
    ("kaynak",             "Kaynak"),
]
PARA_ALANLARI = {"mal_hizmet_toplami", "iskonto", "matrah", "kdv", "tevkifat", "odenecek"}


def csv_yaz(kayitlar, yol):
    with open(yol, "w", newline="", encoding="utf-8-sig") as f:
        y = csv.writer(f, delimiter=";")
        y.writerow([b for _, b in SUTUNLAR])
        for k in kayitlar:
            y.writerow([k.get(a) if k.get(a) is not None else "" for a, _ in SUTUNLAR])


def xlsx_yaz(kayitlar, yol):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Faturalar"

    baslik_dolgu = PatternFill("solid", fgColor="1F4E5F")
    baslik_font = Font(bold=True, color="FFFFFF")
    for i, (_, baslik) in enumerate(SUTUNLAR, start=1):
        h = ws.cell(row=1, column=i, value=baslik)
        h.fill, h.font = baslik_dolgu, baslik_font
        h.alignment = Alignment(horizontal="center", vertical="center")

    for r, kayit in enumerate(kayitlar, start=2):
        for c, (alan, _) in enumerate(SUTUNLAR, start=1):
            h = ws.cell(row=r, column=c, value=kayit.get(alan))
            if alan in PARA_ALANLARI:
                h.number_format = '#,##0.00'
            elif alan in ("satici_vkn", "alici_vkn", "fatura_no", "ettn"):
                h.alignment = Alignment(horizontal="left")
                h.number_format = "@"

    # Toplam satırı
    son = len(kayitlar) + 1
    t = son + 1
    ws.cell(row=t, column=1, value="TOPLAM").font = Font(bold=True)
    for c, (alan, _) in enumerate(SUTUNLAR, start=1):
        if alan in PARA_ALANLARI and kayitlar:
            s = get_column_letter(c)
            h = ws.cell(row=t, column=c, value=f"=SUM({s}2:{s}{son})")
            h.font, h.number_format = Font(bold=True), '#,##0.00'

    genislik = {"kdv_detay": 34, "satici_unvan": 34, "alici_unvan": 34,
                "ettn": 38, "dosya": 26, "kaynak": 34}
    for i, (alan, baslik) in enumerate(SUTUNLAR, start=1):
        ws.column_dimensions[get_column_letter(i)].width = genislik.get(alan, max(12, len(baslik) + 3))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SUTUNLAR))}{son}"
    wb.save(yol)


# ---------------------------------------------------------------------- main

def main(argv=None):
    ap = argparse.ArgumentParser(description="e-Arşiv / e-Fatura ayıklayıcı")
    ap.add_argument("girdiler", nargs="+", help="klasör veya dosya (.eml/.zip/.xml/.pdf)")
    ap.add_argument("-o", "--cikti", default="e-arsiv-faturalari", help="çıktı dosya adı (uzantısız)")
    ap.add_argument("--vkn", help="kendi VKN/TCKN'niz — ALIŞ/SATIŞ yönünü işaretler")
    a = ap.parse_args(argv)

    kayitlar, atlanan, gorulen = [], [], 0

    for kaynak, ad, veri in dosyalari_gez(a.girdiler):
        gorulen += 1
        uzanti = os.path.splitext(ad)[1].lower()
        kayit = None
        if uzanti == ".xml" or veri.lstrip()[:5] == b"<?xml" or veri.lstrip()[:1] == b"<":
            kayit = ubl_ayristir(veri, kaynak, ad)
        if kayit is None and (uzanti == ".pdf" or veri[:4] == b"%PDF"):
            kayit = pdf_ayristir(veri, kaynak, ad)
        if kayit is None:
            atlanan.append(f"{kaynak} / {ad}")
            continue
        if a.vkn:
            if kayit.get("satici_vkn") == a.vkn:
                kayit["yon"] = "SATIŞ"
            elif kayit.get("alici_vkn") == a.vkn:
                kayit["yon"] = "ALIŞ"
            else:
                kayit["yon"] = "?"
        kayitlar.append(kayit)

    # ETTN (yoksa fatura no) ile tekilleştir — aynı fatura hem XML hem PDF gelebilir.
    # XML kaydı PDF'e tercih edilir; XML alanları eksiksizdir.
    kayitlar.sort(key=lambda k: (k.get("format") != "XML",))
    benzersiz, gorulen_anahtar, tekrar = [], set(), 0
    for k in kayitlar:
        anahtar = (k.get("ettn") or "").lower() or f"no:{k.get('fatura_no')}"
        if anahtar in gorulen_anahtar:
            tekrar += 1
            continue
        gorulen_anahtar.add(anahtar)
        benzersiz.append(k)
    benzersiz.sort(key=lambda k: (k.get("tarih") or "", k.get("fatura_no") or ""))

    csv_yol, xlsx_yol = f"{a.cikti}.csv", f"{a.cikti}.xlsx"
    csv_yaz(benzersiz, csv_yol)
    xlsx_yaz(benzersiz, xlsx_yol)

    # ------------------------------------------------------------ özet çıktı
    print(f"\nTaranan dosya      : {gorulen}")
    print(f"Bulunan fatura     : {len(benzersiz)}  (tekrar eden {tekrar} kayıt elendi)")
    if atlanan:
        print(f"Fatura olmayan ek  : {len(atlanan)}")
        for x in atlanan[:8]:
            print(f"    - {x}")
        if len(atlanan) > 8:
            print(f"    ... ve {len(atlanan) - 8} tane daha")

    if benzersiz:
        print(f"\n{'Tarih':<11} {'Fatura No':<18} {'Satıcı':<28} {'Matrah':>11} {'KDV':>10} {'Toplam':>11}")
        print("-" * 94)
        for k in benzersiz:
            print(f"{(k.get('tarih') or '-'):<11} {(k.get('fatura_no') or '-'):<18} "
                  f"{(k.get('satici_unvan') or '-')[:28]:<28} "
                  f"{(k.get('matrah') or 0):>11,.2f} {(k.get('kdv') or 0):>10,.2f} "
                  f"{(k.get('odenecek') or 0):>11,.2f}")
        print("-" * 94)
        print(f"{'TOPLAM':<58} "
              f"{sum(k.get('matrah') or 0 for k in benzersiz):>11,.2f} "
              f"{sum(k.get('kdv') or 0 for k in benzersiz):>10,.2f} "
              f"{sum(k.get('odenecek') or 0 for k in benzersiz):>11,.2f}")

    print(f"\nYazıldı: {xlsx_yol}\n         {csv_yol}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
